"""
excel_writer.py
Gera o relatório Excel com três abas:

  1. Transações  — tabela completa com filtros automáticos
  2. Resumo Mensal — totais por categoria e mês
  3. Gráficos     — gráfico de pizza (gastos por categoria)

Usa openpyxl para formatação completa.
"""

from pathlib import Path
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


# ── Paleta de cores ──────────────────────────────────────────────────────────
COR_HEADER      = "1a1a2e"   # azul escuro quase preto
COR_HEADER_FONT = "FFFFFF"
COR_DEBITO      = "FFF0F0"   # vermelho bem claro
COR_CREDITO     = "F0FFF0"   # verde bem claro
COR_TOTAL       = "E8F0FE"   # azul claro
COR_ZEBRA       = "F8F9FA"   # cinza clarinho

FONTE_BASE = "Arial"
FONTE_TAMANHO = 10


def _borda_fina() -> Border:
    lado = Side(style="thin", color="DDDDDD")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _header_cell(ws, row, col, valor, largura=None):
    """Aplica estilo de cabeçalho em uma célula."""
    cell = ws.cell(row=row, column=col, value=valor)
    cell.font = Font(name=FONTE_BASE, bold=True, color=COR_HEADER_FONT, size=FONTE_TAMANHO)
    cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _borda_fina()
    if largura and ws.column_dimensions[get_column_letter(col)].width < largura:
        ws.column_dimensions[get_column_letter(col)].width = largura
    return cell


def _data_cell(ws, row, col, valor, formato=None, fill=None):
    """Aplica estilo de dados em uma célula."""
    cell = ws.cell(row=row, column=col, value=valor)
    cell.font = Font(name=FONTE_BASE, size=FONTE_TAMANHO)
    cell.border = _borda_fina()
    cell.alignment = Alignment(vertical="center")
    if formato:
        cell.number_format = formato
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


# ── Aba 1: Transações ────────────────────────────────────────────────────────

def _aba_transacoes(wb: Workbook, transacoes: list[dict]) -> None:
    ws = wb.active
    ws.title = "Transações"
    ws.freeze_panes = "A2"  # Congela cabeçalho

    # Cabeçalhos
    colunas = ["Data", "Banco", "Descrição", "Tipo", "Categoria", "Valor (R$)"]
    larguras = [12, 12, 45, 10, 18, 14]
    for col, (nome, larg) in enumerate(zip(colunas, larguras), start=1):
        _header_cell(ws, 1, col, nome, larg)

    ws.row_dimensions[1].height = 22

    # Dados
    for i, t in enumerate(transacoes, start=2):
        zebra = COR_ZEBRA if i % 2 == 0 else None
        cor_linha = COR_DEBITO if t["tipo"] == "debito" else COR_CREDITO

        _data_cell(ws, i, 1, t["data"],        "DD/MM/AAAA", cor_linha)
        _data_cell(ws, i, 2, t["banco"],        None,         cor_linha)
        _data_cell(ws, i, 3, t["descricao"],    None,         cor_linha)

        tipo_str = "Débito" if t["tipo"] == "debito" else "Crédito"
        _data_cell(ws, i, 4, tipo_str,          None,         cor_linha)
        _data_cell(ws, i, 5, t["categoria"],    None,         cor_linha)

        cell_valor = _data_cell(ws, i, 6, t["valor"], '#,##0.00', cor_linha)
        cell_valor.alignment = Alignment(horizontal="right")

        # Negativo em vermelho para débitos
        if t["tipo"] == "debito":
            cell_valor.font = Font(name=FONTE_BASE, size=FONTE_TAMANHO, color="CC0000")

    # Filtro automático
    ws.auto_filter.ref = f"A1:F{len(transacoes)+1}"

    # Linha de total
    linha_total = len(transacoes) + 2
    ws.cell(row=linha_total, column=5, value="TOTAL DÉBITOS").font = Font(bold=True, name=FONTE_BASE)
    total_debitos = sum(t["valor"] for t in transacoes if t["tipo"] == "debito")
    c = ws.cell(row=linha_total, column=6, value=total_debitos)
    c.number_format = '#,##0.00'
    c.font = Font(bold=True, color="CC0000", name=FONTE_BASE)
    c.fill = PatternFill("solid", fgColor=COR_TOTAL)

    linha_credito = linha_total + 1
    ws.cell(row=linha_credito, column=5, value="TOTAL CRÉDITOS").font = Font(bold=True, name=FONTE_BASE)
    total_creditos = sum(t["valor"] for t in transacoes if t["tipo"] == "credito")
    c2 = ws.cell(row=linha_credito, column=6, value=total_creditos)
    c2.number_format = '#,##0.00'
    c2.font = Font(bold=True, color="006600", name=FONTE_BASE)
    c2.fill = PatternFill("solid", fgColor=COR_TOTAL)


# ── Aba 2: Resumo Mensal ─────────────────────────────────────────────────────

def _aba_resumo(wb: Workbook, transacoes: list[dict]) -> None:
    ws = wb.create_sheet("Resumo Mensal")
    ws.freeze_panes = "B2"

    # Agrupa por mês e categoria (apenas débitos para análise de gastos)
    debitos = [t for t in transacoes if t["tipo"] == "debito"]

    meses_set = sorted({t["data"].strftime("%Y-%m") for t in debitos})
    categorias_set = sorted({t["categoria"] for t in debitos})

    if not meses_set:
        ws["A1"] = "Nenhum dado de débito encontrado."
        return

    # Cabeçalho: primeira coluna = Categoria, demais = meses
    _header_cell(ws, 1, 1, "Categoria", 20)
    for col, mes in enumerate(meses_set, start=2):
        # Formata mês: 2025-04 → Abr/2025
        dt = datetime.strptime(mes, "%Y-%m")
        label = dt.strftime("%b/%Y").capitalize()
        _header_cell(ws, 1, col, label, 14)

    col_total = len(meses_set) + 2
    _header_cell(ws, 1, col_total, "Total", 14)

    # Agrupa valores
    dados = defaultdict(lambda: defaultdict(float))
    for t in debitos:
        mes = t["data"].strftime("%Y-%m")
        dados[t["categoria"]][mes] += t["valor"]

    for linha, cat in enumerate(categorias_set, start=2):
        zebra = COR_ZEBRA if linha % 2 == 0 else None
        _data_cell(ws, linha, 1, cat, None, zebra)

        total_cat = 0.0
        for col, mes in enumerate(meses_set, start=2):
            val = dados[cat].get(mes, 0.0)
            total_cat += val
            c = _data_cell(ws, linha, col, val if val else None, '#,##0.00', zebra)
            c.alignment = Alignment(horizontal="right")

        c_total = _data_cell(ws, linha, col_total, total_cat, '#,##0.00', COR_TOTAL)
        c_total.font = Font(bold=True, name=FONTE_BASE, size=FONTE_TAMANHO)
        c_total.alignment = Alignment(horizontal="right")

    # Linha de total geral
    linha_total = len(categorias_set) + 2
    _data_cell(ws, linha_total, 1, "TOTAL GERAL", None, COR_TOTAL).font = Font(
        bold=True, name=FONTE_BASE, size=FONTE_TAMANHO
    )
    for col, mes in enumerate(meses_set, start=2):
        total_mes = sum(dados[cat].get(mes, 0.0) for cat in categorias_set)
        c = _data_cell(ws, linha_total, col, total_mes, '#,##0.00', COR_TOTAL)
        c.font = Font(bold=True, name=FONTE_BASE, size=FONTE_TAMANHO)
        c.alignment = Alignment(horizontal="right")

    ws.auto_filter.ref = f"A1:{get_column_letter(col_total)}{len(categorias_set)+1}"


# ── Aba 3: Gráficos ──────────────────────────────────────────────────────────

def _aba_graficos(wb: Workbook, transacoes: list[dict]) -> None:
    ws = wb.create_sheet("Gráficos")

    # Dados auxiliares para o gráfico (pizza por categoria)
    debitos = [t for t in transacoes if t["tipo"] == "debito"]
    totais = defaultdict(float)
    for t in debitos:
        totais[t["categoria"]] += t["valor"]

    if not totais:
        ws["A1"] = "Nenhum dado disponível para gráfico."
        return

    # Escreve tabela auxiliar de dados
    ws["A1"] = "Categoria"
    ws["B1"] = "Total (R$)"
    ws["A1"].font = Font(bold=True, name=FONTE_BASE)
    ws["B1"].font = Font(bold=True, name=FONTE_BASE)

    categorias_ord = sorted(totais.items(), key=lambda x: x[1], reverse=True)
    for i, (cat, val) in enumerate(categorias_ord, start=2):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=round(val, 2)).number_format = '#,##0.00'

    n = len(categorias_ord)

    # Gráfico de pizza
    grafico = PieChart()
    grafico.title = "Gastos por Categoria"
    grafico.style = 10
    grafico.width = 18
    grafico.height = 14

    dados_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    labels_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    grafico.add_data(dados_ref, titles_from_data=True)
    grafico.set_categories(labels_ref)
    grafico.dataLabels = DataLabelList()
    grafico.dataLabels.showPercent = True
    grafico.dataLabels.showCatName = True

    ws.add_chart(grafico, "D2")


# ── Ponto de entrada ─────────────────────────────────────────────────────────

def gerar_excel(transacoes: list[dict], saida: Path) -> None:
    """
    Gera o arquivo Excel completo com três abas.

    Args:
        transacoes: Lista de dicts com campos: data, banco, descricao,
                    valor, tipo, categoria.
        saida:      Caminho do arquivo .xlsx de saída.
    """
    wb = Workbook()

    _aba_transacoes(wb, transacoes)
    _aba_resumo(wb, transacoes)
    _aba_graficos(wb, transacoes)

    # Metadados do arquivo
    wb.properties.title = "Extrato Bancário — bank-extractor"
    wb.properties.creator = "hevkyr / bank-extractor"

    wb.save(saida)
